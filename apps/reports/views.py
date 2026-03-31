from django.views.generic import TemplateView
from django.views import View
from django.contrib.auth.mixins import LoginRequiredMixin
from django.db.models import Count, Sum
from django.http import JsonResponse, HttpResponse
from datetime import timedelta, date
from apps.donors.models import Donor
from apps.donations.models import Donation
from apps.inventory.models import BloodUnit
from apps.blood_requests.models import BloodRequest
import csv
import io
from openpyxl import Workbook
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib.styles import getSampleStyleSheet

class ReportsDashboardView(LoginRequiredMixin, TemplateView):
    template_name = 'reports/dashboard.html'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        start_date = self.request.GET.get('start_date')
        end_date = self.request.GET.get('end_date')

        # Default to last 30 days if not provided
        if not start_date or not end_date:
            today = date.today()
            start_date = (today - timedelta(days=30)).strftime('%Y-%m-%d')
            end_date = today.strftime('%Y-%m-%d')

        context['start_date'] = start_date
        context['end_date'] = end_date

        # Top 10 Donors Leaderboard
        context['top_donors'] = Donor.objects.annotate(
            donation_count=Count('donations')
        ).order_by('-donation_count')[:10]

        # Critical Stock Summary
        blood_groups = [c[0] for c in Donor.BLOOD_GROUP_CHOICES]
        critical_stock = []
        for bg in blood_groups:
            count = BloodUnit.objects.filter(blood_group=bg, status='available').aggregate(total=Sum('units'))['total'] or 0
            if count < 5:
                critical_stock.append({'group': bg, 'count': count})
        context['critical_stock'] = critical_stock

        return context

class DonationTrendAPI(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        labels = []
        data = []
        today = date.today()
        for i in range(11, -1, -1):
            # Calculate month and year
            month = (today.month - i - 1) % 12 + 1
            year = today.year + (today.month - i - 1) // 12
            
            d = date(year, month, 1)
            labels.append(d.strftime('%b %Y'))
            
            count = Donation.objects.filter(
                donation_date__year=year,
                donation_date__month=month
            ).count()
            data.append(count)
        
        return JsonResponse({'labels': labels, 'data': data})

class BloodGroupDistributionAPI(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        distribution = BloodUnit.objects.filter(status='available').values('blood_group').annotate(
            total_units=Sum('units')
        ).order_by('blood_group')
        
        labels = [item['blood_group'] for item in distribution]
        data = [item['total_units'] for item in distribution]
        
        return JsonResponse({'labels': labels, 'data': data})

class KPISummaryAPI(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        today = date.today()
        start_of_month = today.replace(day=1)
        
        total_donors = Donor.objects.filter(created_at__gte=start_of_month).count()
        total_units = Donation.objects.filter(donation_date__gte=start_of_month).aggregate(total=Sum('units_collected'))['total'] or 0
        total_requests = BloodRequest.objects.filter(created_at__gte=start_of_month).count()
        
        return JsonResponse({
            'total_donors': total_donors,
            'total_units': total_units,
            'total_requests': total_requests
        })

class ExportPDFView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="blood_bank_report.pdf"'
        
        buffer = io.BytesIO()
        doc = SimpleDocTemplate(buffer, pagesize=letter)
        styles = getSampleStyleSheet()
        elements = []
        
        elements.append(Paragraph('Blood Bank Management System - Monthly Report', styles['Title']))
        elements.append(Paragraph(f'Date Generated: {date.today()}', styles['Normal']))
        elements.append(Paragraph('<br/><br/>', styles['Normal']))
        
        # KPI Table
        today = date.today()
        start_of_month = today.replace(day=1)
        kpi_data = [
            ['Metric', 'Current Month Value'],
            ['New Donors', Donor.objects.filter(created_at__gte=start_of_month).count()],
            ['Units Collected', Donation.objects.filter(donation_date__gte=start_of_month).aggregate(total=Sum('units_collected'))['total'] or 0],
            ['New Blood Requests', BloodRequest.objects.filter(created_at__gte=start_of_month).count()],
        ]
        
        kpi_table = Table(kpi_data)
        kpi_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(Paragraph('KPI Summary', styles['Heading2']))
        elements.append(kpi_table)
        
        # Top Donors
        elements.append(Paragraph('<br/><br/>Top 10 Donors', styles['Heading2']))
        top_donors = Donor.objects.annotate(
            donation_count=Count('donations')
        ).order_by('-donation_count')[:10]
        
        donor_data = [['Donor Name', 'Blood Group', 'Total Donations']]
        for donor in top_donors:
            donor_data.append([donor.full_name, donor.blood_group, donor.donation_count])
            
        donor_table = Table(donor_data)
        donor_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.red),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        elements.append(donor_table)
        
        doc.build(elements)
        pdf = buffer.getvalue()
        buffer.close()
        response.write(pdf)
        return response

class ExportExcelView(LoginRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        wb = Workbook()
        
        # Sheet 1: KPIs
        ws1 = wb.active
        ws1.title = "KPI Summary"
        ws1.append(['Metric', 'Value'])
        today = date.today()
        start_of_month = today.replace(day=1)
        ws1.append(['New Donors', Donor.objects.filter(created_at__gte=start_of_month).count()])
        ws1.append(['Units Collected', Donation.objects.filter(donation_date__gte=start_of_month).aggregate(total=Sum('units_collected'))['total'] or 0])
        ws1.append(['New Blood Requests', BloodRequest.objects.filter(created_at__gte=start_of_month).count()])
        
        # Sheet 2: Donors
        ws2 = wb.create_sheet(title="Top Donors")
        ws2.append(['Donor Name', 'Blood Group', 'Total Donations'])
        top_donors = Donor.objects.annotate(
            donation_count=Count('donations')
        ).order_by('-donation_count')[:10]
        for donor in top_donors:
            ws2.append([donor.full_name, donor.blood_group, donor.donation_count])
            
        # Sheet 3: Inventory
        ws3 = wb.create_sheet(title="Current Inventory")
        ws3.append(['Blood Group', 'Available Units'])
        distribution = BloodUnit.objects.filter(status='available').values('blood_group').annotate(
            total_units=Sum('units')
        ).order_by('blood_group')
        for item in distribution:
            ws3.append([item['blood_group'], item['total_units']])
            
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="blood_bank_report.xlsx"'
        wb.save(response)
        return response
